"""Excel report generation with openpyxl.

Produces styled ``.xlsx`` files with coloured headers.
"""

from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from src.core.config import settings


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def generate_excel(df: pd.DataFrame, title: str = "Report") -> str:
    """Write *df* to a timestamped Excel file under ``settings.report_dir``.

    The first row contains bold, white-on-blue headers.

    Args:
        df: Data to write.
        title: Sheet / file title prefix.

    Returns:
        Absolute path of the created ``.xlsx`` file.
    """
    out_dir = Path(settings.report_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"{title}_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, 1):
            if isinstance(value, pd.Period):
                value = str(value)
            ws.cell(row=row_idx + 2, column=col_idx, value=value)

    wb.save(str(path))
    return str(path)
